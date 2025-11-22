
# I guess code will break if number of students present in each exam is different
# Nah it wont lil bro(might give some negligible miscalculations (kt ka))
# Text generation ka logic wapis sheet name manual diya hai
# Form mai saare input ke div ko label karna hai


import builtins

_original_print = builtins.print

def selective_print(*args, **kwargs):
    # Allow printing only if the message contains a keyword
    if any("DEBUG-GRADEREPORT" in str(a) for a in args):
        _original_print(*args, **kwargs)

builtins.print = selective_print








from flask import Flask,render_template,request,Response,send_file,redirect,url_for,session
from io import BytesIO
from openpyxl.styles import Alignment,Font
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import pandas as pd
import datetime
#files import
import c_div
import d_div
import count_kt_c
import count_kt_d
# import c_and_d
import toppers_list
import ILO_analysis
import merge_ILO
import c_div_DLO
import D_div_DLO
import toppers_list_DLO
import format
#Blueprint routes import
from c_div_DLO import routes_c_DLO
from merge_ILO import routes_merging_ILO
from ILO_analysis import routes_ILO
from c_div import routes_c
from d_div import routes_d
# from c_and_d import routes_cd
from firstapp.c_div_DLO import c_div_DLO_analysis
from merge import routes_merging
from merge import merg
from merge_DLO import merg_DLO
from merge_DLO import routes_merging_DLO
import urllib.parse
from Models import db
from dotenv import load_dotenv
import os

import sys
sys.stdout.reconfigure(line_buffering=True)

load_dotenv()  # load .env file

app = Flask('__name__',template_folder='templates')



# Read secrets from environment variables
db_user = os.getenv("DB_USER")
db_pass = urllib.parse.quote_plus(os.getenv("DB_PASS"))
db_host = os.getenv("DB_HOST")
db_name = os.getenv("DB_NAME")
secret_key = os.getenv("SECRET_KEY")

app.config['SQLALCHEMY_DATABASE_URI'] = (
    f'postgresql://{db_user}:{db_pass}@{db_host}/{db_name}'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
app.secret_key=secret_key
app.register_blueprint(routes_merging)
app.register_blueprint(routes_merging_DLO)
app.register_blueprint(routes_merging_ILO)
# app.register_blueprint(routes_cd)
app.register_blueprint(routes_c)
app.register_blueprint(routes_d)
app.register_blueprint(routes_c_DLO)
app.register_blueprint(routes_ILO)
@app.route('/',methods=['GET','POST'])
def index():
    # if request.method=='GET':
    #     return render_template('Form.html')
    # elif request.method=='POST':
    #     sub1=request.form.get('Subject-1')
    #     teacher1=request.form.get('Subject-1-teacher')
    #     sub2 = request.form.get('Subject-2')
    #     teacher2 = request.form.get('Subject-2-teacher')
    #     sub3 = request.form.get('Subject-3')
    #     teacher3 = request.form.get('Subject-3-teacher')
    #     sub4 = request.form.get('Subject-4')
    #     teacher4 = request.form.get('Subject-4-teacher')
    #     sub5 = request.form.get('Subject-5')
    #     teacher5 = request.form.get('Subject-5-teacher')
    #     # return redirect(url_for('convert_excel'))
    #     return render_template('index.html')
    if request.method == 'POST':
        year=request.form['year']
        semester=int(request.form['semester'])
        branch=request.form['branch']
        no_of_divisions = int(request.form.get("divisionCount"))

        prn_start = {}
        prn_end = {}
        for i in range(1,no_of_divisions+1):
            prn_start[f"div{i}_prn_start"] = request.form.get(f'DIV{i}-PRN-start')
            prn_end[f"div{i}_prn_end"] = request.form.get(f'DIV{i}-PRN-end')
        # div1_prn_end = request.form.get('DIV1-PRN-end')
        # div2_prn_start = request.form.get('DIV2-PRN-start')
        # div2_prn_end = request.form.get('DIV2-PRN-end')
        # Collecting the teacher data from the form
        teachers = {}
        for i in range(1,no_of_divisions+1):
            teachers[f"teachers_div{i}"] = [
                # request.form['Subject-1'] this will return keyError if value is not present, use when input is required
                # request.form.get("Subject-1", "default value") this will return none so no error if value is not present can also set default value, used when form/input is dynamic or optional
                {'name': request.form.get(f'Subject-1-teacher-div{i}')},
                {'name': request.form.get(f'Subject-2-teacher-div{i}')},
                {'name': request.form.get(f'Subject-3-teacher-div{i}')},
                {'name': request.form.get(f'Subject-4-teacher-div{i}')},
                {'name': request.form.get(f'Subject-5-teacher-div{i}')},
                {'name': request.form.get(f'DLO-1-teacher-div{i}')},
                {'name': request.form.get(f'DLO-2-teacher-div{i}')},
                {'name': request.form.get(f'DLO-3-teacher-div{i}')},
                {'name': request.form.get(f'DLO-4-teacher-div{i}')},
                {'name': request.form.get(f'DLO-5-teacher-div{i}')},
                {'name': request.form.get(f'DLO-6-teacher-div{i}')},
                {'name': request.form.get(f'ILO-1-teacher-div{i}')},
                {'name': request.form.get(f'ILO-2-teacher-div{i}')},
                {'name': request.form.get(f'ILO-3-teacher-div{i}')},
                {'name': request.form.get(f'ILO-4-teacher-div{i}')},

            ]
        subjects=[
            {'subject': request.form.get('Subject-1')},
            {'subject': request.form.get('Subject-2')},
            {'subject': request.form.get('Subject-3')},
            {'subject': request.form.get('Subject-4')},
            {'subject': request.form.get('Subject-5')},
            {'subject': request.form.get('DLO-1')},
            {'subject': request.form.get('DLO-2')},
            {'subject': request.form.get('DLO-3')},
            {'subject': request.form.get('DLO-4')},
            {'subject': request.form.get('DLO-5')},
            {'subject': request.form.get('DLO-6')},

            {'subject': request.form.get('ILO-1')},
            {'subject': request.form.get('ILO-2')},
            {'subject': request.form.get('ILO-3')},
            {'subject': request.form.get('ILO-4')},

        ]
        # teachers_div2 = [
        #     {'name': request.form.get('Subject-1-teacher-div2'), 'subject': request.form.get('Subject-1')},
        #     {'name': request.form.get('Subject-2-teacher-div2'), 'subject': request.form.get('Subject-2')},
        #     {'name': request.form.get('Subject-3-teacher-div2'), 'subject': request.form.get('Subject-3')},
        #     {'name': request.form.get('Subject-4-teacher-div2'), 'subject': request.form.get('Subject-4')},
        #     {'name': request.form.get('Subject-5-teacher-div2'), 'subject': request.form.get('Subject-5')},
        #     {'name': request.form.get('DLO-1-teacher-div1'), 'subject': request.form.get('DLO-1')},
        #     {'name': request.form.get('DLO-2-teacher-div1'), 'subject': request.form.get('DLO-2')},
        #     {'name': request.form.get('DLO-3-teacher-div1'), 'subject': request.form.get('DLO-3')},
        #
        # ]
        # Storing the teacher data in the session
        for i in range(1,no_of_divisions+1):
            session[f'teachers_div{i}'] = teachers[f"teachers_div{i}"]
        session['subjects'] = subjects
        for i in range(1,no_of_divisions+1):
            session[f'div{i}_prn_start'] = prn_start[f"div{i}_prn_start"]
            session[f'div{i}_prn_end'] = prn_end[f"div{i}_prn_end"]
        session['year']=year
        session['semester'] = semester
        session['branch'] = branch
        session['no_of_divisions'] = no_of_divisions
        return redirect(url_for("convert_excel"))
    return render_template('Form.html')



# Retrieving data from session and redirecting to index.html
# @app.route('/retrieve')
# def retrieve():
#     # teachers_div1 = session.get('teachers_div1', [])
#     # teachers_div2 = session.get('teachers_div2', [])
#
#     # teacher1_name = teachers_div1[0]['name']
#     # teacher1_subject = teachers_div1[0]['subject']
#     # teacher2_name = teachers_div1[1]['name']
#     # teacher2_subject = teachers_div1[1]['subject']
#     # teacher3_name = teachers_div1[2]['name']
#     # teacher3_subject = teachers_div1[2]['subject']
#     # teacher4_name = teachers_div1[3]['name']
#     # teacher4_subject = teachers_div1[3]['subject']
#     # teacher5_name = teachers_div1[4]['name']
#     # teacher5_subject = teachers_div1[4]['subject']
#     return render_template('index.html')
@app.route('/convert_excel',methods=['GET','POST'])
def convert_excel():
    # print("Route Hit")
    # print("Files:", request.files)
    # print("Form:", request.form)
    # if request.method == 'POST':
    #     print("✅ It's a POST request")
    #
    # if 'file' in request.files:
    #     print("✅ File was uploaded")


    semester = session.get('semester')
    # div1_prn_start = session.get('div1_prn_start')
    # div1_prn_end = session.get('div1_prn_end')
    # div2_prn_start = session.get('div2_prn_start')
    # div2_prn_end = session.get('div2_prn_end')
    no_of_divisions = session.get('no_of_divisions')

    prn_start = {}
    prn_end={}
    for i in range(1, no_of_divisions + 1):
        prn_start[f"div{i}_prn_start"] = session.get(f"div{i}_prn_start")
        prn_end[f"div{i}_prn_end"] = session.get(f"div{i}_prn_end")
    
    if request.method == 'POST' and 'file' in request.files and (semester == 5 or semester == 6):
        print("Inside If")
        department = session.get('branch')
        year = session.get('year')
        subjects = session.get('subjects', [])
        subject_1 = subjects[0]['subject']
        subject_2 = subjects[1]['subject']
        subject_3 = subjects[2]['subject']
        subject_4 = subjects[3]['subject']
        DLO_1=subjects[5]['subject']
        DLO_2=subjects[6]['subject']
        DLO_3=subjects[7]['subject']
        file = request.files['file']
        file2=request.files['file2']
        df = pd.read_excel(file)
        DLO1_df=pd.read_excel(file2,sheet_name=0)
        DLO2_df = pd.read_excel(file2, sheet_name=1)
        DLO3_df = pd.read_excel(file2, sheet_name=2)

        final_dataframes_dictionary = {}
        kt_df_dictionary = {}
        appeared_div_dictionary = {}
        pass_div_dictionary = {}
        result_dictionary = {}
        for i in range(1, no_of_divisions + 1):
            # main table dataframe of first sheet
            final_dataframes_dictionary[f"new{i}_df"] = c_div_DLO.c_div_DLO_analysis(df, i, DLO1_df, DLO2_df, DLO3_df, prn_start[f"div{i}_prn_start"],prn_end[f"div{i}_prn_end"])
            # kt dataframe of first sheet
            kt_df_dictionary[f"kt_div{i}_df"], appeared_div_dictionary[f"appeared_div{i}"], pass_div_dictionary[f"pass_div{i}"] = count_kt_d.kt_analysis_d(df, prn_start[f"div{i}_prn_start"],prn_end[f"div{i}_prn_end"])
            # result of all divisions
            result_dictionary[f"result_div{i}"] = round(pass_div_dictionary[f"pass_div{i}"] / appeared_div_dictionary[f"appeared_div{i}"] * 100, 2)

        sheet2_curr, sheet2_prev, difference, sheet2_curr_kt, sheet2_prev_kt, total_students, prev_total_students, curr_perc_result, prev_perc_result, prev_year = merg_DLO(df,DLO1_df, DLO2_df, DLO3_df)

        overall_pass = sum(pass_div_dictionary[f"pass_div{i}"] for i in range(1, no_of_divisions + 1))
        overall_appeared = sum(appeared_div_dictionary[f"appeared_div{i}"] for i in range(1, no_of_divisions + 1))
        overall_result = round(overall_pass / overall_appeared * 100, 2)
        toppers = toppers_list.list_all(df)
        subject_1_toppers = toppers_list.sub1_list(df)
        subject_2_toppers = toppers_list.sub2_list(df)
        subject_3_toppers = toppers_list.sub3_list(df)
        subject_4_toppers = toppers_list.sub4_list(df)
        DLO1_toppers = toppers_list_DLO.DLO_list(df,DLO1_df,"EXAM15")
        DLO2_toppers = toppers_list_DLO.DLO_list(df,DLO2_df,"EXAM15")
        DLO3_toppers = toppers_list_DLO.DLO_list(df,DLO3_df,"EXAM15")

        DLO_df=[DLO1_df,DLO2_df,DLO3_df]
        DLO_toppers=[DLO1_toppers,DLO2_toppers,DLO3_toppers]
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine="openpyxl",mode="w") as writer:
            current_row = 3
            for i in range(1, no_of_divisions + 1):
                final_dataframes_dictionary[f"new{i}_df"].to_excel(writer, sheet_name=f"SEM-{semester}", index=True,startrow=current_row)
                current_row += len(final_dataframes_dictionary[f"new{i}_df"]) + 4
                kt_df_dictionary[f"kt_div{i}_df"].to_excel(writer, sheet_name=f"SEM-{semester}", index=False,startrow=current_row)
                current_row += len(kt_df_dictionary[f"kt_div{i}_df"]) + 4

            sheet2_curr.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=1)
            sheet2_prev.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=6)
            difference.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=11)
            sheet2_curr_kt.to_excel(writer, sheet_name="Comparison", index=False, startrow=23, startcol=1)
            sheet2_prev_kt.to_excel(writer, sheet_name="Comparison", index=False, startrow=23, startcol=10)

            row = 3  # Start at row 3

            toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(toppers) + 4

            subject_1_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(subject_1_toppers) + 4

            subject_2_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(subject_2_toppers) + 4

            subject_3_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(subject_3_toppers) + 4

            subject_4_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(subject_4_toppers) + 4

            empty_df = pd.DataFrame([["N/A", "N/A", "N/A"]], columns=["ROLLNO", "NAME", "MARKS"])

            # DLO1_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            # row += len(DLO1_toppers) + 4
            #
            # DLO2_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            # row += len(DLO2_toppers) + 4
            for i in range(1,4):
                if not DLO_df[i-1].empty:
                    DLO_toppers[i-1].to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
                    row += 7
                else:
                    empty_df.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)


        excel_file.seek(0)
        #formatting the excel file
        wb = load_workbook(excel_file)
        format.set_width_for_all_sheets(wb)
        ws=wb[f"SEM-{semester}"]
        format.first_sheet_formatting(ws,department,semester,no_of_divisions,year,result_dictionary,overall_result,36)

        ws = wb["Comparison"]  # Select the correct sheet
        format.second_sheet_formatting(ws,department,year,total_students,prev_year,prev_total_students, curr_perc_result, prev_perc_result)

        ws = wb['Toppers List']
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border  # Apply border
        ws['B1'] = "Toppers List"
        ws["B1"].font = Font(size=12, bold=True)
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B9'] = f"{subject_1}"
        ws["B9"].font = Font(size=12, bold=True)
        ws["B9"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B16'] = f"{subject_2}"
        ws["B16"].font = Font(size=12, bold=True)
        ws["B16"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B23'] = f"{subject_3}"
        ws["B23"].font = Font(size=12, bold=True)
        ws["B23"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B30'] = f"{subject_4}"
        ws["B30"].font = Font(size=12, bold=True)
        ws["B30"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B37'] = f"{DLO_1}"
        ws["B37"].font = Font(size=12, bold=True)
        ws["B37"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B44'] = f"{DLO_2}"
        ws["B44"].font = Font(size=12, bold=True)
        ws["B44"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B51'] = f"{DLO_3}"
        ws["B51"].font = Font(size=12, bold=True)
        ws["B51"].alignment = Alignment(horizontal="center", vertical="center")

        timestamp = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S") #different name ke liye

        final_excel_file = BytesIO()  # Create a new buffer
        wb.save(final_excel_file)
        final_excel_file.seek(0)  # Reset buffer for further use
        # Send the Excel file as an attachment for download
        return send_file(
            final_excel_file,
            as_attachment=True,
            download_name=f"result-SEM{semester}-{timestamp}.xlsx",  # Name of the file to download
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    elif request.method == 'POST' and 'file' in request.files and (semester == 7 or semester == 8):
        print("Inside elIf sem 7,8")
        department = session.get('branch')
        year = session.get('year')
        subjects = session.get('subjects', [])
        subject_1 = subjects[0]['subject']
        subject_2 = subjects[1]['subject']
        DLO_1=subjects[5]['subject']
        DLO_2=subjects[6]['subject']
        DLO_3=subjects[7]['subject']
        DLO_4=subjects[8]['subject']
        DLO_5=subjects[9]['subject']
        DLO_6=subjects[10]['subject']
        ILO_1=subjects[11]['subject']
        ILO_2 = subjects[12]['subject']
        ILO_3 = subjects[13]['subject']
        ILO_4 = subjects[14]['subject']

        file = request.files['file']
        file2=request.files['file2']
        file3 = request.files['file3']
        df = pd.read_excel(file)

        xls2 = pd.ExcelFile(file2)
        sheet_names = xls2.sheet_names
        sheet_count = len(sheet_names)
        print(f"Sheet names: {sheet_names}")
        print(f"Sheet Count: {sheet_count}")

        DLO1_df = pd.read_excel(file2, sheet_name=0) #This is the first bucket NLP and MV maise se ek
        DLO2_df = pd.read_excel(file2, sheet_name=1) #This mostly will be empty as it is MV
        DLO3_df = pd.read_excel(file2, sheet_name=2) #This is the second bucket of DLO BC and ARVR mai se ek
        DLO4_df = pd.read_excel(file2, sheet_name=3)
        DLO5_df = pd.read_excel(file2, sheet_name=4)
        DLO6_df = pd.read_excel(file2, sheet_name=5)


        ILO1_df = pd.read_excel(file3,sheet_name=0)
        ILO2_df = pd.read_excel(file3, sheet_name=1)
        ILO3_df = pd.read_excel(file3, sheet_name=2)
        ILO4_df = pd.read_excel(file3, sheet_name=3)

        final_dataframes_dictionary = {}
        kt_df_dictionary = {}
        appeared_div_dictionary = {}
        pass_div_dictionary = {}
        result_dictionary = {}
        for i in range(1, no_of_divisions + 1):
            # main table dataframe of first sheet
            final_dataframes_dictionary[f"new{i}_df"] = ILO_analysis.ILO_analysis(df, i, DLO1_df, DLO2_df, DLO3_df,DLO4_df,DLO5_df,DLO6_df, ILO1_df,ILO2_df,ILO3_df,ILO4_df,prn_start[f"div{i}_prn_start"],prn_end[f"div{i}_prn_end"])
            # kt dataframe of first sheet
            kt_df_dictionary[f"kt_div{i}_df"], appeared_div_dictionary[f"appeared_div{i}"], pass_div_dictionary[f"pass_div{i}"] = count_kt_d.kt_analysis_d(df, prn_start[f"div{i}_prn_start"],prn_end[f"div{i}_prn_end"])
            # result of all divisions
            result_dictionary[f"result_div{i}"] = round(pass_div_dictionary[f"pass_div{i}"] / appeared_div_dictionary[f"appeared_div{i}"] * 100, 2)

        sheet2_curr, sheet2_prev, difference, sheet2_curr_kt, sheet2_prev_kt, total_students, prev_total_students, curr_perc_result, prev_perc_result, prev_year = merge_ILO.merg_ILO(df,DLO1_df, DLO2_df, DLO3_df, DLO4_df, DLO5_df, DLO6_df, ILO1_df, ILO2_df, ILO3_df, ILO4_df)

        overall_pass = sum(pass_div_dictionary[f"pass_div{i}"] for i in range(1, no_of_divisions + 1))
        overall_appeared = sum(appeared_div_dictionary[f"appeared_div{i}"] for i in range(1, no_of_divisions + 1))
        overall_result = round(overall_pass / overall_appeared * 100, 2)
        toppers = toppers_list.list_all(df)
        subject_1_toppers = toppers_list.sub1_list(df)
        subject_2_toppers = toppers_list.sub2_list(df)
        DLO_toppers={}
        DLO_df=[DLO1_df,DLO2_df,DLO3_df,DLO4_df,DLO5_df,DLO6_df]
        for i in range(1,7):
            DLO_toppers[f"DLO{i}_toppers"] = toppers_list_DLO.DLO_list(df,DLO_df[i-1],"EXAM9" if i<4 else "EXAM12")
        ILO_toppers={}
        ILO_df=[ILO1_df,ILO2_df,ILO3_df,ILO4_df]
        for i in range(1,5):
            ILO_toppers[f"ILO{i}_toppers"] = toppers_list_DLO.DLO_list(df,ILO_df[i-1],"EXAM15")
        # DLO2_toppers = toppers_list_DLO.DLO_list(df,DLO2_df)
        # DLO3_toppers = toppers_list_DLO.DLO_list(df,DLO3_df)
        # DLO4_toppers = toppers_list_DLO.DLO_list(df,DLO4_df)
        # DLO5_toppers = toppers_list_DLO.DLO_list(df,DLO5_df)
        # DLO6_toppers = toppers_list_DLO.DLO_list(df,DLO6_df)
        # ILO1_toppers = toppers_list_DLO.DLO_list(df,ILO1_df)
        # ILO2_toppers = toppers_list_DLO.DLO_list(df,ILO2_df)
        # ILO3_toppers = toppers_list_DLO.DLO_list(df,ILO3_df)
        # ILO4_toppers = toppers_list_DLO.DLO_list(df,ILO4_df)

        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine="openpyxl",mode="w") as writer:
            current_row = 3
            for i in range(1, no_of_divisions + 1):
                final_dataframes_dictionary[f"new{i}_df"].to_excel(writer, sheet_name=f"SEM-{semester}", index=True,startrow=current_row)
                current_row += len(final_dataframes_dictionary[f"new{i}_df"]) + 4
                kt_df_dictionary[f"kt_div{i}_df"].to_excel(writer, sheet_name=f"SEM-{semester}", index=False,startrow=current_row)
                current_row += len(kt_df_dictionary[f"kt_div{i}_df"]) + 4

            sheet2_curr.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=1)
            sheet2_prev.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=6)
            difference.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=11)
            sheet2_curr_kt.to_excel(writer, sheet_name="Comparison", index=False, startrow=24, startcol=1)
            sheet2_prev_kt.to_excel(writer, sheet_name="Comparison", index=False, startrow=24, startcol=10)

            row = 3  # Start at row 3

            toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(toppers) + 4

            subject_1_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(subject_1_toppers) + 4

            subject_2_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
            row += len(subject_2_toppers) + 4

            empty_df = pd.DataFrame([["N/A", "N/A", "N/A"]], columns=["ROLLNO", "NAME", "MARKS"])
            for i in range(1,7):
                if not DLO_df[i-1].empty:
                    DLO_toppers[f"DLO{i}_toppers"].to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
                    # row += len(DLO_toppers[f"DLO{i}_toppers"]) + 4
                    row += 7
                else:
                    empty_df.to_excel(writer, sheet_name="Toppers List", index=False,startrow=row)
                    # row += len(empty_df) + 4
                    row += 7

            for i in range(1,5):
                if not ILO_df[i-1].empty:
                    ILO_toppers[f"ILO{i}_toppers"].to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
                    # row += len(ILO_toppers[f"ILO{i}_toppers"]) + 4
                    row += 7
                else:
                    empty_df.to_excel(writer, sheet_name="Toppers List", index=False, startrow=row)
                    # row += len(empty_df) + 4
                    row += 7


        excel_file.seek(0)

        # formatting the excel file
        wb = load_workbook(excel_file)
        format.set_width_for_all_sheets(wb)
        ws = wb[f"SEM-{semester}"]
        format.first_sheet_formatting(ws, department, semester, no_of_divisions, year, result_dictionary,overall_result, 46)

        ws = wb["Comparison"]  # Select the correct sheet
        format.second_sheet_formatting(ws, department, year, total_students, prev_year, prev_total_students,
                                       curr_perc_result, prev_perc_result)

        ws = wb['Toppers List']
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border  # Apply border
        ws['B1'] = "Toppers List"
        ws["B1"].font = Font(size=12, bold=True)
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B9'] = f"{subject_1}"
        ws["B9"].font = Font(size=12, bold=True)
        ws["B9"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B16'] = f"{subject_2}"
        ws["B16"].font = Font(size=12, bold=True)
        ws["B16"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B23'] = f"{DLO_1}"
        ws["B23"].font = Font(size=12, bold=True)
        ws["B23"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B30'] = f"{DLO_2}"
        ws["B30"].font = Font(size=12, bold=True)
        ws["B30"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B37'] = f"{DLO_3}"
        ws["B37"].font = Font(size=12, bold=True)
        ws["B37"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B44'] = f"{DLO_4}"
        ws["B44"].font = Font(size=12, bold=True)
        ws["B44"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B51'] = f"{DLO_5}"
        ws["B51"].font = Font(size=12, bold=True)
        ws["B51"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B58'] = f"{DLO_6}"
        ws["B58"].font = Font(size=12, bold=True)
        ws["B58"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B65'] = f"{ILO_1}"
        ws["B65"].font = Font(size=12, bold=True)
        ws["B65"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B72'] = f"{ILO_2}"
        ws["B72"].font = Font(size=12, bold=True)
        ws["B72"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B79'] = f"{ILO_3}"
        ws["B79"].font = Font(size=12, bold=True)
        ws["B79"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B86'] = f"{ILO_4}"
        ws["B86"].font = Font(size=12, bold=True)
        ws["B86"].alignment = Alignment(horizontal="center", vertical="center")

        timestamp = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S") #different name ke liye

        final_excel_file = BytesIO()  # Create a new buffer
        wb.save(final_excel_file)
        final_excel_file.seek(0)  # Reset buffer for further use
        # Send the Excel file as an attachment for download
        return send_file(
            final_excel_file,
            as_attachment=True,
            download_name=f"result-SEM{semester}-{timestamp}.xlsx",  # Name of the file to download
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    elif request.method == 'POST' and 'file' in request.files:
        print("Inside elif")
        # new_df = pd.DataFrame()
        # new2_df=pd.DataFrame()
        department = session.get('branch')
        year=session.get('year')
        subjects = session.get('subjects', [])
        subject_1 = subjects[0]
        subject_2 = subjects[1]
        subject_3 = subjects[2]
        subject_4 = subjects[3]
        subject_5 = subjects[4]
        file = request.files['file']
        df = pd.read_excel(file)

        final_dataframes_dictionary={}
        kt_df_dictionary={}
        appeared_div_dictionary={}
        pass_div_dictionary={}
        result_dictionary = {}
        for i in range(1,no_of_divisions+1):
            #main table dataframe of first sheet
            final_dataframes_dictionary[f"new{i}_df"] = c_div.c_div_analysis(df,i,prn_start[f"div{i}_prn_start"], prn_end[f"div{i}_prn_end"])
            #kt dataframe of first sheet
            kt_df_dictionary[f"kt_div{i}_df"],appeared_div_dictionary[f"appeared_div{i}"],pass_div_dictionary[f"pass_div{i}"] = count_kt_d.kt_analysis_d(df,prn_start[f"div{i}_prn_start"], prn_end[f"div{i}_prn_end"])
            #result of all divisions
            result_dictionary[f"result_div{i}"]=round(pass_div_dictionary[f"pass_div{i}"]/appeared_div_dictionary[f"appeared_div{i}"]*100,2)

        sheet2_curr,sheet2_prev,difference,sheet2_curr_kt,sheet2_prev_kt,total_students,prev_total_students,curr_perc_result,prev_perc_result,prev_year1=merg(df)

        overall_pass = sum(pass_div_dictionary[f"pass_div{i}"] for i in range(1,no_of_divisions+1))
        overall_appeared = sum(appeared_div_dictionary[f"appeared_div{i}"] for i in range(1,no_of_divisions+1))
        overall_result=round(overall_pass/overall_appeared*100,2)
        toppers=toppers_list.list_all(df)
        subject_1_toppers=toppers_list.sub1_list(df)
        subject_2_toppers=toppers_list.sub2_list(df)
        subject_3_toppers=toppers_list.sub3_list(df)
        subject_4_toppers = toppers_list.sub4_list(df)
        subject_5_toppers = toppers_list.sub5_list(df)
        # Create an in-memory bytes buffer
        excel_file = BytesIO()

        # Write to the same sheet with empty rows in between
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            current_row=3
            for i in range(1,no_of_divisions+1):
                final_dataframes_dictionary[f"new{i}_df"].to_excel(writer, sheet_name=f"SEM-{semester}", index=True, startrow=current_row)
                current_row += len(final_dataframes_dictionary[f"new{i}_df"]) +4
                kt_df_dictionary[f"kt_div{i}_df"].to_excel(writer, sheet_name=f"SEM-{semester}", index=False, startrow=current_row)
                current_row += len(kt_df_dictionary[f"kt_div{i}_df"])+4
                # [f"new2_df"].to_excel(writer, sheet_name=f"SEM-{semester}", index=True, startrow=len(new_df) + 14)
                # kt_d_df.to_excel(writer, sheet_name=f"SEM-{semester}", index=False, startrow=len(new_df)+24)

            # cd_df.to_excel(writer, sheet_name="Both Div", index=True, startrow=9)
            sheet2_curr.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=1)
            sheet2_prev.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=6)
            difference.to_excel(writer, sheet_name="Comparison", index=False, startrow=9, startcol=11)
            sheet2_curr_kt.to_excel(writer, sheet_name="Comparison", index=False, startrow=23, startcol=1)
            sheet2_prev_kt.to_excel(writer, sheet_name="Comparison", index=False, startrow=23, startcol=10)


            toppers.to_excel(writer, sheet_name="Toppers List",index=False,startrow=3)
            subject_1_toppers.to_excel(writer, sheet_name="Toppers List",index=False,startrow=10)
            subject_2_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=17)
            subject_3_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=24)
            subject_4_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=31)
            subject_5_toppers.to_excel(writer, sheet_name="Toppers List", index=False, startrow=38)

            # Move to the beginning of the BytesIO buffer
        excel_file.seek(0)

        # formatting the excel file
        wb = load_workbook(excel_file)
        format.set_width_for_all_sheets(wb)
        ws = wb[f"SEM-{semester}"]
        format.first_sheet_formatting(ws, department, semester, no_of_divisions, year, result_dictionary,overall_result, 46)

        ws = wb["Comparison"]  # Select the correct sheet
        format.second_sheet_formatting(ws, department, year, total_students, prev_year1, prev_total_students,curr_perc_result, prev_perc_result)


        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # third sheet formatting
        ws=wb['Toppers List']
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border  # Apply border
        ws['B1']="Toppers List"
        ws["B1"].font = Font(size=12, bold=True)
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B9']=f"{subject_1}"
        ws["B9"].font = Font(size=12, bold=True)
        ws["B9"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B16']=f"{subject_2}"
        ws["B16"].font = Font(size=12, bold=True)
        ws["B16"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B23']=f"{subject_3}"
        ws["B23"].font = Font(size=12, bold=True)
        ws["B23"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B30']=f"{subject_4}"
        ws["B30"].font = Font(size=12, bold=True)
        ws["B30"].alignment = Alignment(horizontal="center", vertical="center")
        ws['B37']=f"{subject_5}"
        ws["B37"].font = Font(size=12, bold=True)
        ws["B37"].alignment = Alignment(horizontal="center", vertical="center")


        final_excel_file = BytesIO()  # Create a new buffer
        wb.save(final_excel_file)
        final_excel_file.seek(0)  # Reset buffer for further use
        # Send the Excel file as an attachment for download
        return send_file(
            final_excel_file,
            as_attachment=True,
            download_name="result.xlsx",  # Name of the file to download
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return render_template('index.html',semester=semester)
if __name__=='__main__':
    app.run(debug=True,use_reloader=False)