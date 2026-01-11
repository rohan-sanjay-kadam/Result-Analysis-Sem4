from openpyxl.styles import Alignment,Font
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def set_width_for_all_sheets(wb):
    # Adjust column width for all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B')
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width  # Set width dynamically



def first_sheet_formatting(ws,department,semester,no_of_divisions,year,result_dictionary,overall_result,s):
    # Apply Borders to All Cells with Data

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border  # Apply border
    ws.merge_cells('B1:K1')
    ws["B1"] = "SIES GRADUATE SCHOOL Of TECHNOLOGY"
    ws["B1"].font = Font(size=30, bold=True)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('B2:K2')
    ws["B2"] = f"DEPARTMENT OF {department}"
    ws["B2"].font = Font(size=13, bold=True)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('B3:K3')
    ws["B3"] = f"Result Analysis -  SEM{semester} "
    ws["B3"].font = Font(size=13, bold=True)
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    start = s  # start 46 for sem7 36 0r 37 for sem5 35 for sem3
    for i in range(1, no_of_divisions + 1):
        row = start + (i - 1) * 2
        ws.merge_cells(f'B{row}:G{row}')
        key = f"result_div{i}"
        ws[f"B{row}"] = f"% RESULT {chr(64 + i)} division OF {year} BATCH: {result_dictionary[key]}%"

    start = start + no_of_divisions * 2
    ws.merge_cells(f'B{start}:G{start}')
    ws[f"B{start}"] = f"% OVERALL RESULT of all divisions OF {year} BATCH: {overall_result}%"

def second_sheet_formatting(ws,department,year,total_students,prev_year1,prev_total_students,curr_perc_result,prev_perc_result):
    ws.merge_cells('A1:Q1')
    ws['A1'] = "SIES GRADUATE SCHOOL OF TECHNOLOGY"
    ws["A1"].font = Font(size=35, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A4:Q4')
    ws['A4'] = f"DEPARTMENT OF {department}"
    ws["A4"].font = Font(size=15, bold=True)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A5:Q5')
    ws['A5'] = "RESULT ANALYSIS"
    ws["A5"].font = Font(size=15, bold=True)
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    # Merge A1:C1 and add text
    for row in ws.iter_rows(min_row=9, max_row=10, min_col=2, max_col=2):
        for cell in row:
            cell.border = thin_border
    ws.merge_cells('B9:B10')
    ws['B9'] = 'Subject'
    for row in ws.iter_rows(min_row=9, max_row=9, min_col=3, max_col=6):  # C=3 to F=6
        for cell in row:
            cell.border = thin_border
    ws.merge_cells('C9:F9')
    ws['C9'] = f"{year}  BATCH PERCENTAGE OF STUDENTS SCORING IN THE RANGE OF (NO. OF STUDENTS = {total_students})"
    ws["C9"].alignment = Alignment(wrapText=True)
    ws.row_dimensions[9].height = 50
    ws.merge_cells('G9:G10')
    ws['G9'] = 'Subject'
    for row in ws.iter_rows(min_row=9, max_row=9, min_col=8, max_col=11):  # H=8 to K=11
        for cell in row:
            cell.border = thin_border
    ws.merge_cells('H9:K9')
    ws[
        'H9'] = f"{prev_year1}  BATCH PERCENTAGE OF STUDENTS SCORING IN THE RANGE OF (NO. OF STUDENTS = {prev_total_students})"
    ws["H9"].alignment = Alignment(wrapText=True)
    for row in ws.iter_rows(min_row=9, max_row=9, min_col=12, max_col=15):  # L=12 to O=15
        for cell in row:
            cell.border = thin_border
    ws.merge_cells('L9:O9')
    ws['L9'] = "Percentage Change (+/-)"
    ws['L9'].alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=24, max_row=24, min_col=2, max_col=9):  # B=2 to I=9
        for cell in row:
            cell.border = thin_border
    ws.merge_cells('B24:I24')
    ws['B24'] = f"{year} Batch Number of Students"
    ws['B24'].alignment = Alignment(horizontal="center", vertical="center")
    # Apply border BEFORE merging
    for row in ws.iter_rows(min_row=24, max_row=24, min_col=11, max_col=18):  # K=11 to R=18
        for cell in row:
            cell.border = thin_border
    #Then merging
    ws.merge_cells('K24:R24')
    ws['K24'] = f"{prev_year1} Batch Number of Students"
    ws['K24'].alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border  # Apply border
    ws.merge_cells('B28:J28')
    ws['B28'] = f"% Result of {year} Batch = {curr_perc_result}%"
    ws.merge_cells('K28:S28')
    ws['K28'] = f"% Result of {prev_year1} Batch = {prev_perc_result}%"
